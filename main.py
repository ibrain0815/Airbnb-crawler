"""
에어비앤비(airbnb.co.kr) 홈페이지 동적 크롤러
브라우저로 열고, 사용자가 원하는 페이지로 이동한 뒤 숙소 목록 수집
Edge 또는 Chrome 지원 (로봇 감지 완화 적용)

실행: python main.py → 버튼이 있는 GUI 창이 뜹니다.
     python main.py --console → 터미널 전용 모드.
"""

import os
import re
import sys
import time

# exe로 실행 시 작업 폴더 = exe 위치, 아니면 스크립트 위치
if getattr(sys, "frozen", False):
    _PROJECT_DIR = os.path.dirname(sys.executable)
    os.chdir(_PROJECT_DIR)  # exe: 작업 폴더를 exe 위치로 (base_library.zip\.wdm 오류 방지)
else:
    _PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))

# ChromeDriver/EdgeDriver 캐시를 exe(또는 스크립트) 폴더로만 사용 (base_library.zip\.wdm 오류 방지)
_WDM_CACHE = os.path.join(_PROJECT_DIR, "wdm_drivers")
os.makedirs(_WDM_CACHE, exist_ok=True)
os.environ["WDM_SSL_VERIFY"] = "0"
# exe 실행 시 WDM_LOCAL 설정 금지: WDM_LOCAL=1이면 webdriver_manager가 sys.path[0]/.wdm 사용 â PyInstaller에서 base_library.zip\.wdm 오류
if not getattr(sys, "frozen", False):
    os.environ["WDM_LOCAL"] = "1"
from datetime import datetime
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# 사용할 브라우저: "edge" | "chrome"
BROWSER = "chrome"

# EdgeDriver 수동 경로 (필요 시 환경변수 EDGE_DRIVER_PATH 또는 여기 지정)
EDGE_DRIVER_PATH = os.environ.get("EDGE_DRIVER_PATH", "")

# Chrome 로봇 감지 우회 (undetected-chromedriver). exe 실행 시에는 비활성화(캐시 경로 오류 방지)
USE_UNDETECTED = False
if BROWSER == "chrome" and not getattr(sys, "frozen", False):
    try:
        import undetected_chromedriver as uc
        USE_UNDETECTED = True
    except ImportError:
        pass

from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
try:
    from webdriver_manager.core.driver_cache import DriverCacheManager
    _WDM_CACHE_MANAGER = DriverCacheManager(_WDM_CACHE)
except Exception:
    _WDM_CACHE_MANAGER = None


# 에어비앤비 크롤링 대상
AIRBNB_HOME_URL = "https://www.airbnb.co.kr/homes"
AIRBNB_BASE = "https://www.airbnb.co.kr"

# 크롤링 최대 페이지 수 (1~3)
MAX_PAGES = 3

# 브라우저/드라이버 (속도·봇 감지 균형)
IMPLICIT_WAIT_SEC = 5
WINDOW_SIZE = (1920, 1080)
# 지연 시간 (봇 감지 우회, 문서 권장 범위 내) - 속도 최적화
DELAY_FIRST_PAGE_SEC = 0.8
DELAY_NEXT_PAGE_AFTER_CLICK_SEC = 0.8
DELAY_BETWEEN_PAGES_SEC = 0.8

# 제공 HTML 구조 기준 (카드 내 data-testid 사용으로 정확·빠름)
LISTING_LINK_SELECTOR = 'a[href*="/rooms/"][aria-labelledby^="title_"]'
# 가격: 가격 전용 스타일만 사용 (속도 개선)
PRICE_ROW_SELECTOR = '[data-testid="price-availability-row"]'  # 카드 찾기용
PRICE_SPAN_STYLE = 'span[style*="pricing-guest-primary-line-unit-price"]'  # ₩159,765 전용
# 평점/리뷰: 동일 카드 내 span[aria-hidden="true"] 중 "4.87 (23)" 형태 또는 "평점 N점, 후기 N개" span
RATING_SPAN_ARIA_HIDDEN = 'span[aria-hidden="true"]'  # 텍스트 예: "5.0 (29)"
RATING_SPAN_CONTAINS = "평점"  # "평점 5.0점(5점 만점), 후기 29개"
# 주소/위치
ADDRESS_SELECTORS = [
    '[data-testid="listing-card-subtitle"]',
    '[data-testid="listing-card-location"]',
]

# 고속 수집: execute_script 1회로 카드 전체 수집 (페이지당 1번 왕복)
_FAST_SCRAPE_SCRIPT = """
var base = "https://www.airbnb.co.kr";
var out = [], seen = {};

// 카드 컨테이너 직접 검색 (더 빠르고 정확함)
var cards = document.querySelectorAll('div[data-testid="card-container"]');

for (var i = 0; i < cards.length; i++) {
  try {
    var card = cards[i];

    // 1. 링크 및 URL 추출
    var link = card.querySelector('a[href*="/rooms/"]');
    if (!link) continue;

    var href = (link.getAttribute("href") || "").trim();
    if (!href || href.indexOf("/rooms/") === -1) continue;
    if (href.charAt(0) === "/") href = base + href;
    if (seen[href]) continue;
    seen[href] = true;

    // 2. 제목 추출
    var title = "";
    var titleEl = card.querySelector('[data-testid="listing-card-title"]');
    if (titleEl) {
      title = (titleEl.textContent || "").trim();
    }
    if (!title) title = "(숙소명 없음)";

    // 3. 가격 추출 (취소선 제외, 최저가 선택)
    var price = "";
    var allPrices = [];

    // span.u174bpcy 우선 확인
    var priceSpans = card.querySelectorAll('span.u174bpcy');
    for (var j = 0; j < priceSpans.length; j++) {
      var ps = priceSpans[j];
      var style = window.getComputedStyle(ps);
      var isStrike = (style && style.textDecorationLine && style.textDecorationLine.indexOf("line-through") !== -1);
      var cls = ps.className || "";
      if (cls.indexOf("sjwpj0z") !== -1) isStrike = true;

      if (!isStrike) {
        var txt = (ps.textContent || "").trim();
        if (/^₩[\\d,]+$/.test(txt)) {
          allPrices.push(txt);
        }
      }
    }

    // 없으면 span.u1opajno 확인
    if (allPrices.length === 0) {
      var discSpans = card.querySelectorAll('span.u1opajno');
      for (var j = 0; j < discSpans.length; j++) {
        var txt = (discSpans[j].textContent || "").trim();
        if (/^₩[\\d,]+$/.test(txt)) {
          allPrices.push(txt);
        }
      }
    }

    // 최저가 선택
    if (allPrices.length > 0) {
      var minPrice = allPrices[0];
      var minVal = parseInt(minPrice.replace(/[^0-9]/g, "")) || 0;
      for (var j = 1; j < allPrices.length; j++) {
        var val = parseInt(allPrices[j].replace(/[^0-9]/g, "")) || 0;
        if (val > 0 && val < minVal) {
          minVal = val;
          minPrice = allPrices[j];
        }
      }
      price = minPrice;
    }

    // 4. 평점/리뷰 추출
    var rating = "";

    // span.r4a59j5 aria-label 확인
    var ratingSpan = card.querySelector('span.r4a59j5');
    if (ratingSpan) {
      var ariaLabel = (ratingSpan.getAttribute("aria-label") || "").trim();
      var ratingText = (ratingSpan.textContent || "").trim();

      // "신규 숙소" 체크
      if ((ariaLabel.indexOf("신규") !== -1 && ariaLabel.indexOf("숙소") !== -1) ||
          (ratingText.indexOf("신규") !== -1 && ratingText.indexOf("숙소") !== -1)) {
        rating = "신규숙소";
      } else {
        // aria-label에서 평점 정보 추출
        var m = ariaLabel.match(/평점\\s*([\\d.]+).*?후기\\s*(\\d+)/);
        if (m) {
          rating = m[1] + " (" + m[2] + ")";
        } else if (ariaLabel) {
          rating = ariaLabel;
        } else if (ratingText) {
          rating = ratingText;
        }
      }
    }

    // 없으면 다른 방법으로 시도
    if (!rating) {
      var spans = card.querySelectorAll('span.a8jt5op');
      for (var j = 0; j < spans.length; j++) {
        var t = (spans[j].textContent || "").trim();
        if (t.indexOf("신규") !== -1 && t.indexOf("숙소") !== -1) {
          rating = "신규숙소";
          break;
        }
        if (/^\\d\\.?\\d*\\s*\\(\\d+\\)/.test(t)) {
          rating = t;
          break;
        }
      }
    }

    // 5. 주소/위치 추출
    var address = "";
    var subtitles = card.querySelectorAll('[data-testid="listing-card-subtitle"]');
    var parts = [];
    for (var j = 0; j < subtitles.length; j++) {
      var pt = (subtitles[j].textContent || "").trim();
      if (pt) parts.push(pt);
    }
    if (parts.length > 0) {
      address = parts.join(" | ");
    }

    // 결과 추가
    out.push({
      title: title,
      price: price,
      rating: rating,
      address: address,
      link: href
    });

  } catch(e) {
    // 개별 카드 오류는 무시하고 계속 진행
    continue;
  }
}

return out;
"""

# User-Agent
USER_AGENT_EDGE = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0"
)
USER_AGENT_CHROME = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
)

# 로봇 감지 우회 CDP 스크립트
STEALTH_SCRIPT = """
Object.defineProperty(navigator, 'webdriver', { get: function() { return undefined; } });
Object.defineProperty(navigator, 'languages', { get: function() { return ['ko-KR', 'ko', 'en-US', 'en']; } });
window.chrome = window.chrome || { runtime: {} };
"""


def _apply_stealth(driver: WebDriver) -> None:
    try:
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": STEALTH_SCRIPT})
    except Exception:
        pass


def create_driver(headless: bool = False):
    """Edge 또는 Chrome 드라이버 생성"""
    if BROWSER == "edge":
        options = EdgeOptions()
        if headless:
            options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_argument("--lang=ko-KR")
        options.add_argument(f"--user-agent={USER_AGENT_EDGE}")
        options.add_experimental_option("useAutomationExtension", False)

        if EDGE_DRIVER_PATH and os.path.isfile(EDGE_DRIVER_PATH):
            service = EdgeService(executable_path=EDGE_DRIVER_PATH)
            driver = webdriver.Edge(service=service, options=options)
        else:
            try:
                if _WDM_CACHE_MANAGER is not None:
                    service = EdgeService(EdgeChromiumDriverManager(cache_manager=_WDM_CACHE_MANAGER).install())
                else:
                    service = EdgeService(EdgeChromiumDriverManager().install())
                driver = webdriver.Edge(service=service, options=options)
            except Exception as e:
                if "Connection" in str(type(e).__name__) or "getaddrinfo" in str(e).lower():
                    print("EdgeDriver 자동 다운로드 실패. Selenium 내장 방식으로 시도...")
                driver = webdriver.Edge(options=options)
        _apply_stealth(driver)
    else:
        if USE_UNDETECTED:
            options = uc.ChromeOptions()
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_argument("--lang=ko-KR")
            options.add_argument(f"--user-agent={USER_AGENT_CHROME}")
            if headless:
                options.add_argument("--headless=new")
            driver = uc.Chrome(options=options, use_subprocess=True)
        else:
            options = ChromeOptions()
            if headless:
                options.add_argument("--headless")
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_argument("--lang=ko-KR")
            options.add_argument(f"--user-agent={USER_AGENT_CHROME}")
            options.add_experimental_option("useAutomationExtension", False)
            if _WDM_CACHE_MANAGER is not None:
                service = ChromeService(ChromeDriverManager(cache_manager=_WDM_CACHE_MANAGER).install())
            else:
                service = ChromeService(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            _apply_stealth(driver)
    driver.implicitly_wait(IMPLICIT_WAIT_SEC)
    driver.set_window_size(1280, 900)
    return driver


def open_airbnb_page(driver: WebDriver) -> bool:
    """에어비앤비 홈(숙소 목록) 페이지로 이동"""
    driver.get(AIRBNB_HOME_URL)
    time.sleep(0.8)
    return True


def accept_cookie_if_any(driver: WebDriver) -> None:
    """쿠키/동의 팝업 버튼이 있으면 클릭"""
    selectors = [
        'button[data-testid="accept-cookie-banner"]',
        'button:contains("동의"), button:contains("Accept")',
        'a[href*="cookie"]',
        '[aria-label*="동의"], [aria-label*="Accept"]',
        'button[class*="accept"], button[class*="agree"]',
    ]
    for sel in selectors:
        try:
            if ":contains" in sel:
                continue
            btn = driver.find_element(By.CSS_SELECTOR, sel)
            if btn.is_displayed():
                btn.click()
                time.sleep(0.5)
                break
        except Exception:
            continue


# 평점/리뷰 텍스트 패턴: "4.87 (23)" 또는 "5.0 (29)"
RATING_PATTERN = re.compile(r"^\d\.?\d*\s*\(\d+\)\s*$")
# 카드 내 평점 형태 검색용 (텍스트 일부에서 추출)
RATING_IN_TEXT = re.compile(r"\d\.?\d*\s*\(\d+\)")


def _get_card_container(link_el):
    """링크에 속한 카드. listing-card → card-container → price-availability-row 포함 조상."""
    try:
        el = link_el
        for _ in range(20):
            try:
                parent = el.find_element(By.XPATH, "..")
            except Exception:
                break
            el = parent
            try:
                tid = (el.get_attribute("data-testid") or "").strip()
                if tid == "listing-card" or tid == "card-container":
                    return el
            except Exception:
                pass
        el = link_el
        best = None
        for _ in range(15):
            try:
                parent = el.find_element(By.XPATH, "..")
            except Exception:
                break
            el = parent
            try:
                rows = el.find_elements(By.CSS_SELECTOR, PRICE_ROW_SELECTOR)
                if not rows:
                    continue
                # 카드 하나만 포함하는 조상 우선: 내부에 PRICE_ROW_SELECTOR가 1개인 것이 가장 좁은 카드
                if len(rows) == 1:
                    best = el
                elif best is None:
                    best = el
            except Exception:
                continue
        return best
    except Exception:
        pass
    return None


# 가격: aria-label/텍스트에서 총액(첫 번째 ₩숫자)만 추출
_PRICE_TOTAL_PATTERN = re.compile(r"총액\s*(₩[\d,]+)")
_PRICE_AMOUNT_ONLY = re.compile(r"^₩[\d,]+$")
_PRICE_FIRST_AMOUNT = re.compile(r"₩[\d,]+")


def _price_only_total(raw: str) -> str:
    """총액/실제 표시 금액만. '총액 ₩1,155,213, 원래 요금 ₩1,264,913'→'₩1,155,213'. '₩679,000 · 5박, 원래 요금 ₩920,957'→'₩679,000'(원래 요금 앞 금액)."""
    if not raw or "₩" not in raw:
        return ""
    s = (raw or "").strip().rstrip(",").strip()
    m = _PRICE_TOTAL_PATTERN.search(s)
    if m:
        return m.group(1).strip()
    if "원래 요금" in s:
        before = s.split("원래 요금")[0].strip()
        m = _PRICE_FIRST_AMOUNT.search(before)
        return m.group(0).strip() if m else ""
    if _PRICE_AMOUNT_ONLY.match(s):
        return s
    m = _PRICE_FIRST_AMOUNT.search(s)
    return m.group(0) if m else ""


def _is_strikethrough(span) -> bool:
    """취소선(strikethrough) 스타일 확인"""
    try:
        # 클래스 확인 (sjwpj0z는 취소선 가격)
        cls = span.get_attribute("class") or ""
        if "sjwpj0z" in cls:
            return True
        # 스타일 확인
        style = span.get_attribute("style") or ""
        if "line-through" in style:
            return True
        # CSS computed style 확인
        decoration = span.value_of_css_property("text-decoration-line") or ""
        if "line-through" in decoration:
            return True
    except Exception:
        pass
    return False


def _parse_price_value(price_str: str) -> int:
    """가격 문자열(₩1,234,567)을 숫자로 변환"""
    if not price_str:
        return 0
    import re
    num = re.sub(r"[^0-9]", "", price_str)
    return int(num) if num else 0


def _get_price_from_card(card) -> str:
    """카드 내 가격 추출. 여러 가격 중 가장 낮은 가격 반환 (취소선 가격 제외)."""
    if not card:
        return ""
    all_prices = []

    try:
        # 1) span.u174bpcy 우선 확인 (가장 흔한 가격 표시)
        try:
            spans = card.find_elements(By.CSS_SELECTOR, "span.u174bpcy")
            for span in spans:
                if not _is_strikethrough(span):
                    txt = (span.text or "").strip()
                    if _PRICE_AMOUNT_ONLY.match(txt):
                        all_prices.append(txt)
        except Exception:
            pass

        # 2) 없으면 span.u1opajno 확인 (할인 가격)
        if not all_prices:
            try:
                spans = card.find_elements(By.CSS_SELECTOR, "span.u1opajno")
                for span in spans:
                    if not _is_strikethrough(span):
                        txt = (span.text or "").strip()
                        if _PRICE_AMOUNT_ONLY.match(txt):
                            all_prices.append(txt)
            except Exception:
                pass

        # 최저가 반환
        if all_prices:
            min_price = all_prices[0]
            min_val = _parse_price_value(min_price)
            for price in all_prices[1:]:
                val = _parse_price_value(price)
                if val > 0 and val < min_val:
                    min_val = val
                    min_price = price
            return min_price

    except Exception:
        pass
    return ""


def _get_rating_from_card(card) -> str:
    """카드 내 평점/리뷰 추출. span.r4a59j5 aria-label 우선 사용."""
    if not card:
        return ""
    try:
        # 1) span.r4a59j5 aria-label 우선 확인
        try:
            span = card.find_element(By.CSS_SELECTOR, "span.r4a59j5")
            # aria-label 확인
            aria_label = (span.get_attribute("aria-label") or "").strip()
            text = (span.text or "").strip()

            # 신규 숙소 체크
            if ("신규" in aria_label and "숙소" in aria_label) or ("신규" in text and "숙소" in text):
                return "신규숙소"

            # aria-label에서 평점 정보 추출: "평점 5.0점(5점 만점), 후기 12개"
            if aria_label:
                m = re.search(r"평점\s*([\d.]+).*?후기\s*(\d+)", aria_label)
                if m:
                    return f"{m.group(1)} ({m.group(2)})"
                # aria-label 그대로 반환
                return aria_label

            # textContent 사용
            if text:
                return text
        except Exception:
            pass

        # 2) span.a8jt5op 확인 (백업)
        try:
            spans = card.find_elements(By.CSS_SELECTOR, "span.a8jt5op")
            for span in spans:
                t = (span.text or "").strip()
                if "신규" in t and "숙소" in t:
                    return "신규숙소"
                if RATING_PATTERN.match(t):
                    return t
        except Exception:
            pass

    except Exception:
        pass
    return ""


def _get_address_from_element(container) -> str:
    """컨테이너 내 주소/위치. listing-card-subtitle 텍스트 합침."""
    if not container:
        return ""
    try:
        parts = []
        for el in container.find_elements(By.CSS_SELECTOR, '[data-testid="listing-card-subtitle"]'):
            t = (el.text or "").strip()
            if t:
                parts.append(t)
        if parts:
            return " | ".join(parts)
        for sel in ADDRESS_SELECTORS:
            try:
                el = container.find_element(By.CSS_SELECTOR, sel)
                t = (el.text or "").strip()
                if t:
                    return t
            except Exception:
                continue
    except Exception:
        pass
    return ""


def _get_address_from_card(card) -> str:
    """카드 내 주소/위치 추출. listing-card-subtitle → listing-card-location 순."""
    return _get_address_from_element(card)


def _get_address_near_link(link_el) -> str:
    """링크의 조상에서 주소/위치 요소 탐색 (카드가 좁을 때 누락 방지)."""
    try:
        el = link_el
        for _ in range(20):
            try:
                parent = el.find_element(By.XPATH, "..")
            except Exception:
                break
            el = parent
            t = _get_address_from_element(el)
            if t:
                return t
    except Exception:
        pass
    return ""


def get_airbnb_listings(driver: WebDriver) -> list[dict]:
    """현재 페이지 숙소 수집. 고속 수집(execute_script 1회) 우선, 실패 시 SELECTORS fallback."""
    # 고속 수집: 1회 스크립트로 전체 카드 반환 (페이지당 1번 왕복)
    try:
        raw = driver.execute_script(_FAST_SCRAPE_SCRIPT)
        if raw and isinstance(raw, list) and len(raw) > 0:
            return [
                {
                    "title": (x.get("title") or "").strip() or "(숙소명 없음)",
                    "price": (x.get("price") or "").strip().rstrip(",").strip(),
                    "rating": (x.get("rating") or "").strip(),
                    "address": (x.get("address") or "").strip(),
                    "link": (x.get("link") or "").strip(),
                }
                for x in raw
                if (x.get("link") or "").strip()
            ]
    except Exception:
        pass

    # Fallback: 카드 컨테이너 직접 검색
    results = []
    seen_links = set()

    try:
        cards = driver.find_elements(By.CSS_SELECTOR, 'div[data-testid="card-container"]')
    except Exception:
        cards = []

    for card in cards:
        try:
            # 1. 링크 추출
            link_el = card.find_element(By.CSS_SELECTOR, 'a[href*="/rooms/"]')
            href = (link_el.get_attribute("href") or "").strip()
            if not href or "/rooms/" not in href:
                continue
            if href.startswith("/"):
                href = AIRBNB_BASE.rstrip("/") + href
            if href in seen_links:
                continue
            seen_links.add(href)

            # 2. 제목 추출
            title = ""
            try:
                title_el = card.find_element(By.CSS_SELECTOR, '[data-testid="listing-card-title"]')
                title = (title_el.text or "").strip()
            except Exception:
                pass
            if not title:
                title = "(숙소명 없음)"

            # 3. 가격 추출
            price = _get_price_from_card(card)

            # 4. 평점 추출
            rating = _get_rating_from_card(card)

            # 5. 주소 추출
            address = _get_address_from_element(card)

            results.append({
                "title": title,
                "price": price or "",
                "rating": rating or "",
                "address": address or "",
                "link": href,
            })
        except Exception:
            continue

    return results


def save_listings_to_excel(listings: list[dict], filepath: str | None = None) -> str:
    """수집된 숙소 목록을 엑셀 파일로 저장. 저장 경로 반환."""
    if filepath is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"airbnb_listings_{timestamp}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "숙소 목록"
    headers = ["번호", "숙소명", "가격", "평점/후기", "주소/위치", "링크"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row, item in enumerate(listings, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=item.get("title", ""))
        ws.cell(row=row, column=3, value=item.get("price", ""))
        ws.cell(row=row, column=4, value=item.get("rating", ""))
        ws.cell(row=row, column=5, value=item.get("address", ""))
        ws.cell(row=row, column=6, value=item.get("link", ""))
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 60
    # 첫 행 고정 (스크롤 시 헤더 고정)
    ws.freeze_panes = "A2"
    # 모든 셀 가운데 정렬
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    last_row = len(listings) + 1
    for row in range(1, last_row + 1):
        for col in range(1, 7):
            ws.cell(row=row, column=col).alignment = center
    wb.save(filepath)
    return filepath


def go_to_next_page(driver: WebDriver) -> bool:
    """다음 페이지로 이동 (다음 버튼 클릭 또는 URL 변경). 성공 시 True."""
    # 에어비앤비: 다음 버튼/링크 선택자 (여러 후보 시도)
    next_selectors = [
        ('css', 'a[aria-label="다음"]'),
        ('css', 'a[aria-label="Next"]'),
        ('css', '[data-testid="pagination-next"]'),
        ('css', 'a[href*="items_offset"]'),
        ('xpath', '//a[contains(@aria-label,"다음") or contains(@aria-label,"Next")]'),
        ('xpath', '//button[contains(text(),"다음") or contains(text(),"Next")]'),
    ]
    for kind, sel in next_selectors:
        try:
            by = By.CSS_SELECTOR if kind == "css" else By.XPATH
            el = driver.find_element(by, sel)
            if el.is_displayed():
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
                time.sleep(0.2)
                el.click()
                time.sleep(DELAY_NEXT_PAGE_AFTER_CLICK_SEC)
                return True
        except Exception:
            continue
    return False


def main():
    driver = None
    try:
        print(f"{'Edge' if BROWSER == 'edge' else 'Chrome'} 브라우저 시작..." + (" (로봇 감지 우회)" if USE_UNDETECTED else ""))
        driver = create_driver(headless=False)
        time.sleep(0.5)

        print("에어비앤비 홈페이지 열기...")
        open_airbnb_page(driver)

        input("\n>>> 검색/필터 후 원하는 목록이 보이면, 여기 터미널에서 엔터를 누르면 크롤링을 시작합니다.\n")

        current_url = driver.current_url
        print(f"현재 페이지: {current_url[:90]}{'...' if len(current_url) > 90 else ''}\n")

        accept_cookie_if_any(driver)
        time.sleep(0.2)

        all_listings = []
        seen_links = set()

        for page_num in range(MAX_PAGES):
            print(f"숙소 목록 수집 중... ({page_num + 1}/{MAX_PAGES}페이지)")
            page_listings = get_airbnb_listings(driver)
            for item in page_listings:
                link = item.get("link", "")
                if link and link not in seen_links:
                    seen_links.add(link)
                    all_listings.append(item)
            if page_num < MAX_PAGES - 1 and not go_to_next_page(driver):
                print(f"다음 페이지가 없어 {page_num + 1}페이지까지 수집했습니다.")
                break
            time.sleep(DELAY_BETWEEN_PAGES_SEC)

        if not all_listings:
            print("수집된 숙소가 없습니다. 페이지가 완전히 로드된 뒤 다시 엔터를 눌러 보세요.")
        else:
            print(f"\n수집된 숙소: {len(all_listings)}개 (최대 {MAX_PAGES}페이지)\n")

            # 상세 페이지에서 좌표(위도·경도) 추가 수집
            try:
                print("숙소 좌표(위도·경도) 수집 중... (상세 페이지 방문)")
                added = enrich_listings_with_coordinates(driver, all_listings)
                print(f"좌표가 추가된 숙소: {added}개\n")
            except Exception as e:
                print(f"좌표 수집 중 오류 발생: {e}\n")

            for i, item in enumerate(all_listings, 1):
                print(f"[{i}] {item['title']}")
                if item.get("price"):
                    print(f"    가격: {item['price']}")
                if item.get("rating"):
                    print(f"    평점/후기: {item['rating']}")
                if item.get("address"):
                    print(f"    주소/위치: {item['address']}")
                print(f"    링크: {item['link']}\n")

            excel_path = save_listings_to_excel(all_listings)
            print(f"엑셀 저장 완료: {excel_path}")

        input("엔터를 누르면 브라우저를 종료합니다...")
    finally:
        if driver:
            driver.quit()
            print("브라우저 종료됨.")


if __name__ == "__main__":
    # 크롤링 버튼이 있는 GUI 실행 (콘솔만 쓰려면: python main.py --console)
    if "--console" not in sys.argv:
        try:
            import gui_app
            gui_app.main()
        except Exception as e:
            print(f"\nGUI 오류: {e}")
            import traceback
            traceback.print_exc()
            input("\n엔터를 누르면 종료합니다...")
    else:
        try:
            main()
        except Exception as e:
            print(f"\n오류 발생: {e}")
            import traceback
            traceback.print_exc()
            input("\n엔터를 누르면 종료합니다...")
